import streamlit as st
import pandas as pd
from datetime import datetime, time
import io
import base64
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client, Client

# 페이지 기본 설정
st.set_page_config(page_title="통합 급여·초과근무·연차 관리 시스템", layout="wide")

# -------------------------------------------------------------------
# Supabase 클라우드 DB 연결 설정 (Secrets 참조)
# -------------------------------------------------------------------
try:
    url: str = st.secrets["https://supabase.com/dashboard/project/vumwmqbgmpygqpiofcwc"]
    key: str = st.secrets["eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZ1bXdtcWJnbXB5Z3FwaW9mY3djIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODc4MTAzMTMsImV4cCI6MjEwMzM4NjMxM30.xv4hMvNkgyjneD7hEvOZ4pZV-irIcNVJug2BK7_m5E8"]
    supabase: Client = create_client(url, key)
except Exception as e:
    st.error("Supabase 연결 실패! Streamlit Secrets에 SUPABASE_URL과 SUPABASE_KEY가 정상 설정되었는지 확인이 필요하다.")
    st.stop()

def truncate_ten(value):
    return int(value // 10) * 10

# 세션 내 로고 이미지 관리
if 'logo_b64' not in st.session_state:
    st.session_state.logo_b64 = ""

st.title("🏢 통합 급여·초과근무·연차 관리 시스템")

# 사이드바: 회사 로고 업로드 기능
with st.sidebar:
    st.header("🖼️ 회사 로고 설정")
    uploaded_logo = st.file_uploader("모든 문서에 적용할 로고 이미지 (PNG, JPG)", type=['png', 'jpg', 'jpeg'])
    if uploaded_logo is not None:
        bytes_data = uploaded_logo.getvalue()
        st.session_state.logo_b64 = base64.b64encode(bytes_data).decode()
        st.image(uploaded_logo, caption="등록된 회사 로고", use_container_width=True)
    elif st.session_state.logo_b64:
        st.info("💡 기존에 등록된 로고가 적용 중이다.")

# 탭 구성 (tab1 ~ tab9)
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs([
    "👥 직원 등록 및 정보 관리", 
    "📝 초과/휴일근무 신청", 
    "✅ 초과근무 수행 입력 & 요약표", 
    "🌴 연차 관리 & 전 직원 요약표",
    "🖨️ 연차 신청서 출력",
    "📊 통합 급여대장 (수정 및 엑셀)", 
    "📄 개별 급여명세서 인쇄",
    "🖨️ 통합 급여대장 인쇄",
    "📑 월별 급여대장 총괄표"
])

# -------------------------------------------------------------------
# TAB 1: 직원 등록 및 정보 관리
# -------------------------------------------------------------------
with tab1:
    st.header("1. 신규 직원 등록")
    with st.form("employee_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            emp_id = st.text_input("사번")
            emp_name = st.text_input("이름")
            birth_date = st.text_input("생년월일 (예: 1980-01-01)", value="1980-01-01")
            dept = st.text_input("부서")
            position = st.text_input("직위", value="주임")
            hobong = st.text_input("호봉", value="1호봉")
            total_leave = st.number_input("연간 총 연차 부여일수", min_value=0.0, value=15.0, step=0.5)
        with col2:
            base_salary = st.number_input("기본급 (원)", min_value=0, value=2500000, step=100000)
            hourly_wage = st.number_input("통상시급 (원)", min_value=0, value=12000, step=500)
            family_allowance = st.number_input("가족수당 (원)", min_value=0, value=50000, step=10000)
            non_taxable = st.number_input("비과세 (원)", min_value=0, value=100000, step=10000)
            other_allowance = st.number_input("기타수당 (원)", min_value=0, value=0, step=10000)
            other_deduction = st.number_input("기타공제 (원)", min_value=0, value=0, step=10000)

            st.write("**4대보험 가입 여부 선택**")
            is_national = st.checkbox("국민연금 가입", value=True)
            is_health = st.checkbox("건강/장기요양보험 가입", value=True)
            is_employment = st.checkbox("고용보험 가입", value=True)
            is_industrial = st.checkbox("산재보험 가입", value=True)
            
        submit_emp = st.form_submit_button("직원 DB 등록")
        
        if submit_emp:
            if emp_id and emp_name:
                emp_data = {
                    "emp_id": emp_id, "emp_name": emp_name, "birth_date": birth_date, "dept": dept,
                    "position": position, "hobong": hobong, "base_salary": base_salary,
                    "hourly_wage": hourly_wage, "family_allowance": family_allowance,
                    "non_taxable": non_taxable, "other_allowance": other_allowance, "other_deduction": other_deduction,
                    "is_national": 1 if is_national else 0, "is_health": 1 if is_health else 0,
                    "is_employment": 1 if is_employment else 0, "is_industrial": 1 if is_industrial else 0,
                    "total_annual_leave": total_leave
                }
                res = supabase.table("employees").insert(emp_data).execute()
                if res.data:
                    st.success(f"{emp_name} ({position}) 직원이 Supabase DB에 정상 등록되었다.")
                else:
                    st.error("직원 등록 중 오류가 발생했거나 이미 존재하는 사번이다.")
            else:
                st.error("사번과 이름을 모두 입력해야 한다.")

    st.divider()
    st.header("2. 누적 직원 데이터 조회 및 편집")
    try:
    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()
    except Exception as e:
    st.warning("⚠️ Supabase에서 직원 데이터를 불러오지 못했다. Secrets의 API Key 설정 또는 DB 상태를 확인해 주어야 한다.")
    df_emp = pd.DataFrame()
    
    if not df_emp.empty:
        edited_df = st.data_editor(df_emp, use_container_width=True, num_rows="dynamic")
        if st.button("수정 데이터 DB 저장"):
            for _, row in edited_df.iterrows():
                supabase.table("employees").upsert(row.to_dict()).execute()
            st.success("직원 데이터 수정사항이 Supabase DB에 반영되었다.")
            st.rerun()

# -------------------------------------------------------------------
# TAB 2: 초과근무 사전 신청
# -------------------------------------------------------------------
with tab2:
    st.header("1. 초과근무 / 휴일근무 사전 신청")
    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()

    if df_emp.empty:
        st.warning("먼저 '직원 등록 및 정보 관리' 탭에서 직원을 등록해야 한다.")
    else:
        emp_list = df_emp['emp_name'] + " (" + df_emp['position'] + " / " + df_emp['emp_id'] + ")"
        selected_emp_str = st.selectbox("직원 선택", emp_list, key="ot_emp")
        selected_emp_id = selected_emp_str.split("/")[-1].replace(")", "").strip()
        emp_info = df_emp[df_emp['emp_id'] == selected_emp_id].iloc[0]

        work_date = st.date_input("근무 예정 일자", datetime.now())
        work_type = st.radio("근무 구분", ["평일 초과근무 (18:00 이후)", "휴일근무"])
        
        col1, col2 = st.columns(2)
        with col1:
            start_time = st.time_input("예정 시작 시간", time(18, 0) if work_type == "평일 초과근무 (18:00 이후)" else time(9, 0))
        with col2:
            end_time = st.time_input("예정 종료 시간", time(20, 0))

        reason = st.text_area("신청 사유")

        start_dt = datetime.combine(work_date, start_time)
        end_dt = datetime.combine(work_date, end_time)
        duration_hours = (end_dt - start_dt).total_seconds() / 3600

        if duration_hours < 0:
            st.error("종료 시간은 시작 시간보다 빨라야 한다.")
        else:
            multiplier = 1.5
            raw_pay = duration_hours * emp_info['hourly_wage'] * multiplier
            estimated_pay = truncate_ten(raw_pay)

            st.info(f"💡 예상 근무시간: **{duration_hours:.1f}시간**")

            if st.button("사전 신청서 제출"):
                ot_data = {
                    "apply_dt": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "emp_id": emp_info['emp_id'], "emp_name": emp_info['emp_name'],
                    "dept": emp_info['dept'], "position": emp_info['position'],
                    "work_date": str(work_date), "work_type": work_type,
                    "start_time": str(start_time), "end_time": str(end_time),
                    "duration_hours": duration_hours, "estimated_pay": estimated_pay,
                    "act_start_time": str(start_time), "act_end_time": str(end_time),
                    "actual_duration_hours": duration_hours, "actual_pay": estimated_pay,
                    "status": "신청", "reason": reason, "act_reason": ""
                }
                supabase.table("overtime_records").insert(ot_data).execute()
                st.success("초과근무 신청 내역이 Supabase DB에 등록되었다.")

# -------------------------------------------------------------------
# TAB 3: 실제 수행 입력 & 삭제 기능 & 월별 승인 요약표
# -------------------------------------------------------------------
with tab3:
    st.header("✅ 실제 초과근무 수행 내역 입력 & 월별 승인 요약표")
    
    ot_res = supabase.table("overtime_records").select("*").order("id", desc=True).execute()
    df_ot = pd.DataFrame(ot_res.data) if ot_res.data else pd.DataFrame()

    if df_ot.empty:
        st.info("등록된 초과근무 신청 내역이 없다.")
    else:
        col_ot1, col_ot2 = st.columns([3, 1])
        with col_ot1:
            ot_options = [f"ID {r['id']} | [{r['status']}] [{r['work_date']}] {r['emp_name']} ({r['work_type']}) - 사전: {float(r['duration_hours']):.1f}h" for _, r in df_ot.iterrows()]
            selected_ot_idx = st.selectbox("처리 및 출력할 초과근무 내역 선택", range(len(ot_options)), format_func=lambda x: ot_options[x])
            target_ot = df_ot.iloc[selected_ot_idx]
        
        with col_ot2:
            st.write("**🗑️ 선택 내역 삭제**")
            if st.button("해당 초과근무 내역 삭제", type="primary", key="del_ot_btn"):
                supabase.table("overtime_records").delete().eq("id", int(target_ot['id'])).execute()
                st.success(f"ID {target_ot['id']} 초과근무 내역이 정상적으로 삭제되었다.")
                st.rerun()

        emp_s_res = supabase.table("employees").select("*").eq("emp_id", target_ot['emp_id']).execute()
        df_emp_single = pd.DataFrame(emp_s_res.data) if emp_s_res.data else pd.DataFrame()
        hourly_w = df_emp_single.iloc[0]['hourly_wage'] if not df_emp_single.empty else 12000

        act_s_val = target_ot.get('act_start_time', target_ot['start_time'])
        act_e_val = target_ot.get('act_end_time', target_ot['end_time'])
        if pd.isna(act_s_val) or not act_s_val: act_s_val = target_ot['start_time']
        if pd.isna(act_e_val) or not act_e_val: act_e_val = target_ot['end_time']

        st.subheader(f"✏️ 실제 수행 시간 및 승인 상태 변경: {target_ot['emp_name']} ({target_ot['work_date']})")
        
        with st.form("actual_ot_form"):
            col_a1, col_a2 = st.columns(2)
            with col_a1:
                st.write(f"**사전 신청 정보**")
                st.write(f"- 근무 구분: {target_ot['work_type']}")
                st.write(f"- 신청 시간: {target_ot['start_time']} ~ {target_ot['end_time']} ({float(target_ot['duration_hours']):.1f}시간)")
                st.write(f"- 신청 사유: {target_ot['reason']}")
            
            with col_a2:
                st.write(f"**실제 수행 근무시간 & 업무 내용 입력**")
                try:
                    init_s_time = datetime.strptime(str(act_s_val)[:8], "%H:%M:%S").time()
                    init_e_time = datetime.strptime(str(act_e_val)[:8], "%H:%M:%S").time()
                except:
                    init_s_time = time(18, 0)
                    init_e_time = time(20, 0)

                act_s_time = st.time_input("실제 시작 시간", init_s_time)
                act_e_time = st.time_input("실제 종료 시간", init_e_time)

                dummy_date = datetime.now().date()
                s_dt = datetime.combine(dummy_date, act_s_time)
                e_dt = datetime.combine(dummy_date, act_e_time)
                calculated_act_hours = max(0.0, (e_dt - s_dt).total_seconds() / 3600)

                st.metric(label="실제 인정 시간", value=f"{calculated_act_hours:.1f} 시간")
                act_pay = truncate_ten(calculated_act_hours * hourly_w * 1.5)
                
                status_choice = st.selectbox("승인 상태", ["승인", "신청", "반려"], index=["승인", "신청", "반려"].index(target_ot['status']) if target_ot['status'] in ["승인", "신청", "반려"] else 0)
                act_reason_input = st.text_input("실제 업무 수행 내용 / 확인 메모", value=target_ot['act_reason'] if pd.notna(target_ot['act_reason']) else '')

            submit_act = st.form_submit_button("실제 수행 내역 저장 및 승인 반영")

            if submit_act:
                update_data = {
                    "act_start_time": str(act_s_time), "act_end_time": str(act_e_time),
                    "actual_duration_hours": calculated_act_hours, "actual_pay": act_pay,
                    "status": status_choice, "act_reason": act_reason_input
                }
                supabase.table("overtime_records").update(update_data).eq("id", int(target_ot['id'])).execute()
                st.success(f"[{target_ot['emp_name']}] 직원의 수행 내역 및 승인 상태({status_choice})가 반영되었다.")
                st.rerun()

        latest_res = supabase.table("overtime_records").select("*").eq("id", int(target_ot['id'])).execute()
        target_ot_latest = latest_res.data[0] if latest_res.data else target_ot

        st.divider()
        st.subheader("🖨️ 초과근무 신청 및 확인서 인쇄")

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 35px; float: left;">' if st.session_state.logo_b64 else ''
        act_reason_disp = target_ot_latest['act_reason'] if pd.notna(target_ot_latest['act_reason']) and target_ot_latest['act_reason'] != "" else "입력된 실제 수행 내용 없음"

        ot_confirm_template = f"""
        <div style="text-align: right; margin-bottom: 10px;">
            <button onclick="window.print()" style="padding: 8px 16px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px;">🖨️ 해당 서식 인쇄하기</button>
        </div>
        <div style="border: 2px solid #000; padding: 30px; font-family: 'Malgun Gothic', sans-serif; max-width: 680px; margin: auto; background: #fff;">
            {logo_html}
            <div style="display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px; clear: both;">
                <h2 style="margin: 0; padding-top: 15px; font-size: 22px; text-decoration: underline;">초 과 근 무 신 청 및 확 인 서</h2>
                <table style="border-collapse: collapse; text-align: center; font-size: 12px; width: 210px;" border="1">
                    <tr style="height: 20px; background-color: #f2f2f2;">
                        <th rowspan="2" style="width: 25px; background-color: #e6e6e6;">결<br>재</th>
                        <th style="width: 60px;">담 당</th>
                        <th style="width: 60px;">대 리</th>
                        <th style="width: 65px;">센터장</th>
                    </tr>
                    <tr style="height: 50px;">
                        <td></td><td></td><td></td>
                    </tr>
                </table>
            </div>

            <table style="width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 13px;" border="1">
                <tr style="height: 38px;">
                    <th style="padding: 6px; background: #f9f9f9; width: 20%;">성 명</th>
                    <td style="padding: 6px; width: 30%;">{target_ot_latest['emp_name']} ({target_ot_latest['position']})</td>
                    <th style="padding: 6px; background: #f9f9f9; width: 20%;">소 속</th>
                    <td style="padding: 6px; width: 30%;">{target_ot_latest['dept']}</td>
                </tr>
                <tr style="height: 38px;">
                    <th style="padding: 6px; background: #f9f9f9;">근무구분</th>
                    <td style="padding: 6px;" colspan="3">{target_ot_latest['work_type']} (최종 상태: {target_ot_latest['status']})</td>
                </tr>
                <tr style="height: 38px;">
                    <th style="padding: 6px; background: #f9f9f9;">사전 신청일시</th>
                    <td style="padding: 6px;" colspan="3">{target_ot_latest['work_date']} ({target_ot_latest['start_time']} ~ {target_ot_latest['end_time']}) / {float(target_ot_latest['duration_hours']):.1f}시간</td>
                </tr>
                <tr style="height: 40px; background-color: #ffffcc;">
                    <th style="padding: 6px; background: #fff2cc;">실제 수행 인정</th>
                    <td style="padding: 6px;" colspan="3"><b>실제 근무시간: {target_ot_latest['act_start_time']} ~ {target_ot_latest['act_end_time']} ({float(target_ot_latest['actual_duration_hours']):.1f} 시간)</b></td>
                </tr>
                <tr>
                    <th style="padding: 6px; background: #f9f9f9;">사유 및 업무내용</th>
                    <td style="padding: 10px; height: 70px; vertical-align: top;" colspan="3">
                        <b>[신청 사유]</b> {target_ot_latest['reason']}<br>
                        <b>[실제 수행 내용]</b> {act_reason_disp}
                    </td>
                </tr>
            </table>

            <p style="text-align: center; margin-top: 35px; font-size: 14px;">위와 같이 초과근무를 신청하고 실제 수행 내역을 확인합니다.</p>
            <p style="text-align: center; margin-top: 10px; font-size: 13px;">{target_ot_latest['work_date'][:4]}년 {target_ot_latest['work_date'][5:7]}월 {target_ot_latest['work_date'][8:10]}일</p>
            
            <p style="text-align: right; margin-top: 30px; font-size: 14px; font-weight: bold; padding-right: 10px;">
                신청 및 확인인: {target_ot_latest['emp_name']} (인)
            </p>
        </div>
        """
        st.components.v1.html(ot_confirm_template, height=560, scrolling=True)

        st.divider()
        st.subheader("📊 월별 승인 초과근무 집계 요약표")
        
        current_year = datetime.now().year
        c_y, c_m = st.columns(2)
        with c_y:
            sel_year = st.selectbox("조회 연도 선택", range(current_year - 2, current_year + 3), index=2, key="ot_year_sel")
        with c_m:
            sel_month = st.selectbox("조회 월 선택", range(1, 13), index=datetime.now().month - 1, key="ot_month_sel")
        
        filter_month = f"{sel_year}-{sel_month:02d}"

        ot_m_res = supabase.table("overtime_records").select("*").like("work_date", f"{filter_month}%").eq("status", "승인").execute()
        df_ot_month = pd.DataFrame(ot_m_res.data) if ot_m_res.data else pd.DataFrame()

        if df_ot_month.empty:
            st.info(f"💡 [{filter_month}] 승인 완료된 초과근무 내역이 없다.")
        else:
            summary_ot = df_ot_month.groupby(['emp_id', 'emp_name', 'dept', 'position', 'status']).agg(
                승인_건수=('id', 'count'),
                총_인정시간_h=('actual_duration_hours', 'sum')
            ).reset_index()
            summary_ot['총_인정시간_h'] = summary_ot['총_인정시간_h'].round(1)
            
            st.write(f"**[{filter_month}] 최종 승인된 직원별 초과근무 인정 시간 현황**")
            st.dataframe(summary_ot, use_container_width=True)

# -------------------------------------------------------------------
# TAB 4: 개인별 연차 관리 & 전 직원 연차 요약표
# -------------------------------------------------------------------
with tab4:
    st.header("🌴 개인별 연차 관리 & 전 직원 연차 요약표")
    
    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()

    if df_emp.empty:
        st.warning("등록된 직원이 없다.")
    else:
        col_l1, col_l2 = st.columns([1, 1])
        
        with col_l1:
            st.subheader("1. 연차/휴가 신청서 작성")
            emp_leave_list = df_emp['emp_name'] + " (" + df_emp['position'] + " / " + df_emp['emp_id'] + ")"
            selected_l_emp = st.selectbox("직원 선택", emp_leave_list, key="leave_emp_select")
            selected_l_id = selected_l_emp.split("/")[-1].replace(")", "").strip()
            l_emp_info = df_emp[df_emp['emp_id'] == selected_l_id].iloc[0]

            leave_type = st.selectbox("휴가 종류", ["연차 (1일)", "오전반차 (0.5일)", "오후반차 (0.5일)", "병가", "경조휴가", "특별휴가"])
            
            l_start_date = st.date_input("휴가 시작일", datetime.now(), key="l_s_date")
            l_end_date = st.date_input("휴가 종료일", datetime.now(), key="l_e_date")
            
            if "반차" in leave_type:
                used_days = 0.5
            else:
                used_days = float((l_end_date - l_start_date).days + 1)

            leave_reason = st.text_area("휴가 사유", key="l_reason")

            if st.button("연차 신청서 제출 및 DB 저장"):
                leave_data = {
                    "apply_dt": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "emp_id": l_emp_info['emp_id'], "emp_name": l_emp_info['emp_name'],
                    "dept": l_emp_info['dept'], "position": l_emp_info['position'],
                    "leave_type": leave_type, "start_date": str(l_start_date),
                    "end_date": str(l_end_date), "used_days": used_days, "reason": leave_reason
                }
                supabase.table("leave_records").insert(leave_data).execute()
                st.success("연차 신청 내역이 Supabase DB에 저장되었다.")
                st.rerun()

        with col_l2:
            st.subheader("2. 개인별 연차 현황 요약 및 삭제")
            l_res = supabase.table("leave_records").select("*").eq("emp_id", l_emp_info['emp_id']).order("id", desc=True).execute()
            df_leave_all = pd.DataFrame(l_res.data) if l_res.data else pd.DataFrame()

            used_annual = df_leave_all[df_leave_all['leave_type'].str.contains("연차|반차", na=False)]['used_days'].sum() if not df_leave_all.empty else 0.0
            total_annual = l_emp_info['total_annual_leave']
            remaining_annual = total_annual - used_annual

            m1, m2, m3 = st.columns(3)
            m1.metric("총 부여 연차", f"{total_annual} 일")
            m2.metric("사용 연차", f"{used_annual} 일")
            m3.metric("잔여 연차", f"{remaining_annual} 일")

            st.write(f"**[{l_emp_info['emp_name']}] 개인 신청 이력**")
            if not df_leave_all.empty:
                st.dataframe(df_leave_all[['id', 'apply_dt', 'leave_type', 'start_date', 'end_date', 'used_days', 'reason']], use_container_width=True)

                st.divider()
                st.write("**🗑️ 연차 신청 내역 삭제**")
                del_leave_options = [f"ID {r['id']} | [{r['start_date']}] {r['leave_type']} ({r['used_days']}일)" for _, r in df_leave_all.iterrows()]
                selected_del_str = st.selectbox("삭제할 내역 선택", del_leave_options, key="del_leave_sel")
                selected_del_id = int(selected_del_str.split("|")[0].replace("ID", "").strip())

                if st.button("선택한 연차 내역 삭제", type="primary"):
                    supabase.table("leave_records").delete().eq("id", selected_del_id).execute()
                    st.success("해당 연차 내역이 정상적으로 삭제되었다.")
                    st.rerun()

        st.divider()
        st.header("📋 센터 등록 직원 전체 연차 내역 요약표")
        
        all_l_res = supabase.table("leave_records").select("*").execute()
        df_all_leaves = pd.DataFrame(all_l_res.data) if all_l_res.data else pd.DataFrame()

        summary_rows = []
        for idx, emp_row in df_emp.iterrows():
            emp_l_records = df_all_leaves[df_all_leaves['emp_id'] == emp_row['emp_id']] if not df_all_leaves.empty else pd.DataFrame()
            u_annual = emp_l_records[emp_l_records['leave_type'].str.contains("연차|반차", na=False)]['used_days'].sum() if not emp_l_records.empty else 0.0
            tot_annual = emp_row['total_annual_leave']
            rem_annual = tot_annual - u_annual
            
            summary_rows.append({
                "사번": emp_row['emp_id'], "이름": emp_row['emp_name'], "부서": emp_row['dept'], "직위": emp_row['position'],
                "총 부여 연차": tot_annual, "사용 연차": u_annual, "잔여 연차": rem_annual,
                "사용률 (%)": round((u_annual / tot_annual * 100), 1) if tot_annual > 0 else 0.0
            })

        df_summary_all = pd.DataFrame(summary_rows)

        output_leave = io.BytesIO()
        with pd.ExcelWriter(output_leave, engine='openpyxl') as writer:
            df_summary_all.to_excel(writer, index=False, sheet_name="전직원_연차_요약")
            worksheet = writer.sheets["전직원_연차_요약"]

            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True)
            body_font = Font(name="맑은 고딕", size=10)

            thin_border = Border(
                left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
            )
            align_center = Alignment(horizontal='center', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = align_center
                    else:
                        cell.font = body_font
                        cell.alignment = align_right if cell.column in [5, 6, 7, 8] else align_center

            for col in worksheet.columns:
                max_len = max(sum(2 if ord(c) > 127 else 1 for c in str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        excel_leave_data = output_leave.getvalue()

        st.download_button(
            label="📥 전 직원 연차 내역 요약표 엑셀 다운로드 (.xlsx)",
            data=excel_leave_data, file_name=f"전직원_연차요약_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.dataframe(df_summary_all, use_container_width=True)

# -------------------------------------------------------------------
# TAB 5: 연차 신청서 독립 출력 탭
# -------------------------------------------------------------------
with tab5:
    st.header("🖨️ 휴가 (연차) 신청서 인쇄")
    
    l_records_res = supabase.table("leave_records").select("*").order("id", desc=True).execute()
    df_leave_records = pd.DataFrame(l_records_res.data) if l_records_res.data else pd.DataFrame()

    if df_leave_records.empty:
        st.info("등록된 연차/휴가 신청 내역이 없다.")
    else:
        leave_options = [f"[{r['start_date']}] {r['emp_name']} {r['position']} - {r['leave_type']}" for _, r in df_leave_records.iterrows()]
        selected_l_index = st.selectbox("출력할 연차 신청서 선택", range(len(leave_options)), format_func=lambda x: leave_options[x])
        target_l = df_leave_records.iloc[selected_l_index]

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 35px; float: left;">' if st.session_state.logo_b64 else ''

        leave_template = f"""
        <div style="text-align: right; margin-bottom: 10px;">
            <button onclick="window.print()" style="padding: 8px 16px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px;">🖨️ 해당 서식 인쇄하기</button>
        </div>
        <div style="border: 2px solid #000; padding: 30px; font-family: 'Malgun Gothic', sans-serif; max-width: 680px; margin: auto; background: #fff;">
            {logo_html}
            <div style="display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px; clear: both;">
                <h2 style="margin: 0; padding-top: 15px; font-size: 24px; text-decoration: underline;">휴 가 (연 차) 신 청 서</h2>
                <table style="border-collapse: collapse; text-align: center; font-size: 12px; width: 210px;" border="1">
                    <tr style="height: 20px; background-color: #f2f2f2;">
                        <th rowspan="2" style="width: 25px; background-color: #e6e6e6;">결<br>재</th>
                        <th style="width: 60px;">담 당</th>
                        <th style="width: 60px;">대 리</th>
                        <th style="width: 65px;">센터장</th>
                    </tr>
                    <tr style="height: 50px;">
                        <td></td><td></td><td></td>
                    </tr>
                </table>
            </div>

            <table style="width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 14px;" border="1">
                <tr style="height: 40px;">
                    <th style="padding: 8px; background: #f9f9f9; width: 20%;">성 명</th>
                    <td style="padding: 8px; width: 30%;">{target_l['emp_name']} ({target_l['position']})</td>
                    <th style="padding: 8px; background: #f9f9f9; width: 20%;">소 속</th>
                    <td style="padding: 8px; width: 30%;">{target_l['dept']}</td>
                </tr>
                <tr style="height: 40px;">
                    <th style="padding: 8px; background: #f9f9f9;">휴가구분</th>
                    <td style="padding: 8px;" colspan="3">{target_l['leave_type']} (사용일수: {target_l['used_days']}일)</td>
                </tr>
                <tr style="height: 40px;">
                    <th style="padding: 8px; background: #f9f9f9;">휴가기간</th>
                    <td style="padding: 8px;" colspan="3">{target_l['start_date']} ~ {target_l['end_date']}</td>
                </tr>
                <tr>
                    <th style="padding: 8px; background: #f9f9f9;">휴가사유</th>
                    <td style="padding: 12px; height: 80px; vertical-align: top;" colspan="3">{target_l['reason']}</td>
                </tr>
            </table>

            <p style="text-align: center; margin-top: 50px; font-size: 15px;">위와 같이 휴가(연차)를 신청합니다.</p>
            <p style="text-align: center; margin-top: 15px; font-size: 13px;">{target_l['apply_dt'][:10]}</p>
            
            <p style="text-align: right; margin-top: 40px; font-size: 15px; font-weight: bold; padding-right: 10px;">
                신청인: {target_l['emp_name']} (인)
            </p>
        </div>
        """
        st.components.v1.html(leave_template, height=560, scrolling=True)

# -------------------------------------------------------------------
# TAB 6: 통합 급여대장 (수정 및 엑셀)
# -------------------------------------------------------------------
with tab6:
    st.header("📊 통합 급여대장 (수정 및 엑셀)")
    
    pay_date = st.date_input("지급일 선택", datetime.now(), key="payroll_date")
    pay_month = pay_date.strftime("%Y-%m")

    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()

    ot_res = supabase.table("overtime_records").select("*").eq("status", "승인").execute()
    df_ot = pd.DataFrame(ot_res.data) if ot_res.data else pd.DataFrame()

    adj_res = supabase.table("monthly_payroll_adjust").select("*").eq("pay_month", pay_month).execute()
    df_adjust = pd.DataFrame(adj_res.data) if adj_res.data else pd.DataFrame()

    if df_emp.empty:
        st.warning("등록된 직원이 없다.")
    else:
        calculated_rows = []
        no = 1
        
        for idx, emp in df_emp.iterrows():
            adj_match = df_adjust[df_adjust['emp_id'] == emp['emp_id']] if not df_adjust.empty else pd.DataFrame()

            emp_ot = df_ot[(df_ot['emp_id'] == emp['emp_id']) & (df_ot['work_date'].str.startswith(pay_month))] if not df_ot.empty else pd.DataFrame()
            calculated_ot_pay = int(emp_ot['actual_pay'].sum()) if not emp_ot.empty else 0

            if not adj_match.empty:
                adj = adj_match.iloc[0]
                base = adj['base_salary']
                ot_pay = adj['ot_pay'] if adj['ot_pay'] > 0 else calculated_ot_pay
                family = adj['family_allowance']
                non_tax = adj['non_taxable']
                other_allow = adj['other_allowance']
                emp_national = adj['national_pension']
                emp_health = adj['health_insurance']
                emp_longterm = adj['longterm_care']
                emp_employment = adj['employment_insurance']
                emp_income_tax = adj['income_tax']
                emp_local_tax = adj['local_tax']
                other_deduct = adj['other_deduction']
            else:
                ot_pay = calculated_ot_pay
                base = emp['base_salary']
                family = emp['family_allowance']
                non_tax = emp['non_taxable']
                other_allow = emp['other_allowance']
                other_deduct = emp['other_deduction']

                total_gross_calc = truncate_ten(base + ot_pay + family + non_tax + other_allow)
                taxable_gross_calc = total_gross_calc - non_tax

                emp_national = truncate_ten(taxable_gross_calc * 0.0475) if emp.get('is_national', 1) == 1 else 0
                emp_health = truncate_ten(taxable_gross_calc * 0.03595) if emp.get('is_health', 1) == 1 else 0
                emp_longterm = truncate_ten(emp_health * 0.1295) if emp.get('is_health', 1) == 1 else 0
                emp_employment = truncate_ten(taxable_gross_calc * 0.0090) if emp.get('is_employment', 1) == 1 else 0
                emp_income_tax = truncate_ten(taxable_gross_calc * 0.03)
                emp_local_tax = truncate_ten(emp_income_tax * 0.10)

            tot_g = base + ot_pay + family + non_tax + other_allow
            taxable_gross = tot_g - non_tax

            biz_national = truncate_ten(taxable_gross * 0.0475) if emp.get('is_national', 1) == 1 else 0
            biz_health = truncate_ten(taxable_gross * 0.03595) if emp.get('is_health', 1) == 1 else 0
            biz_longterm = truncate_ten(biz_health * 0.1295) if emp.get('is_health', 1) == 1 else 0
            biz_employment = truncate_ten(taxable_gross * 0.0115) if emp.get('is_employment', 1) == 1 else 0
            biz_industrial = truncate_ten(taxable_gross * 0.0726) if emp.get('is_industrial', 1) == 1 else 0
            retirement_accrual = truncate_ten(tot_g / 12)

            calculated_rows.append({
                "No": no, "사번": emp['emp_id'], "이름": emp['emp_name'], "생년월일": emp['birth_date'], "부서": emp['dept'], "직위": emp['position'], "호봉": emp['hobong'],
                "기본급": base, "초과수당(승인)": ot_pay, "가족수당": family, "비과세": non_tax, "기타수당": other_allow,
                "국민연금(본인)": emp_national, "건강보험(본인)": emp_health, "장기요양(본인)": emp_longterm, "고용보험(본인)": emp_employment,
                "소득세": emp_income_tax, "지방소득세": emp_local_tax, "기타공제": other_deduct,
                "국민연금(사업자)": biz_national, "건강보험(사업자)": biz_health, "장기요양(사업자)": biz_longterm, "고용보험(사업자)": biz_employment, "산재보험(사업자)": biz_industrial,
                "퇴직적립금": retirement_accrual
            })
            no += 1

        df_calc = pd.DataFrame(calculated_rows)

        st.subheader(f"✏️ {pay_month} 급여대장 항목별 수정 편집기")
        st.info("💡 공제·수당 수치를 수정한 후 아래 [수정사항 명세서 반영 저장] 버튼을 누르면 개별 급여명세서에 적용된다.")
        
        edited_payroll = st.data_editor(df_calc, use_container_width=True)

        if st.button("💾 수정사항 개별 급여명세서에 연동 저장"):
            for idx, r in edited_payroll.iterrows():
                pay_adj_data = {
                    "pay_month": pay_month, "emp_id": r['사번'], "base_salary": int(r['기본급']),
                    "ot_pay": int(r['초과수당(승인)']), "family_allowance": int(r['가족수당']),
                    "non_taxable": int(r['비과세']), "other_allowance": int(r['기타수당']),
                    "national_pension": int(r['국민연금(본인)']), "health_insurance": int(r['건강보험(본인)']),
                    "longterm_care": int(r['장기요양(본인)']), "employment_insurance": int(r['고용보험(본인)']),
                    "income_tax": int(r['소득세']), "local_tax": int(r['지방소득세']), "other_deduction": int(r['기타공제'])
                }
                supabase.table("monthly_payroll_adjust").upsert(pay_adj_data).execute()
            st.success(f"{pay_month} 급여대장 수정 수치가 Supabase DB에 저장 및 연동 완료되었다.")

        payroll_html_rows = ""
        sum_base = sum_ot = sum_family = sum_nontax = sum_gross = 0
        sum_nat = sum_hea = sum_long = sum_emp = sum_inc = sum_loc = sum_other_d = sum_deduct_tot = sum_net = 0
        sum_b_nat = sum_b_hea = sum_b_long = sum_b_emp = sum_b_ind = sum_b_tot = sum_retire = 0

        for idx, row in edited_payroll.iterrows():
            total_gross = row['기본급'] + row['초과수당(승인)'] + row['가족수당'] + row['비과세'] + row['기타수당']
            emp_deduction_total = row['국민연금(본인)'] + row['건강보험(본인)'] + row['장기요양(본인)'] + row['고용보험(본인)'] + row['소득세'] + row['지방소득세'] + row['기타공제']
            net_pay = total_gross - emp_deduction_total
            biz_deduction_total = row['국민연금(사업자)'] + row['건강보험(사업자)'] + row['장기요양(사업자)'] + row['고용보험(사업자)'] + row['산재보험(사업자)']

            sum_base += row['기본급']; sum_ot += row['초과수당(승인)']; sum_family += row['가족수당']; sum_nontax += row['비과세']; sum_gross += total_gross
            sum_nat += row['국민연금(본인)']; sum_hea += row['건강보험(본인)']; sum_long += row['장기요양(본인)']; sum_emp += row['고용보험(본인)']
            sum_inc += row['소득세']; sum_loc += row['지방소득세']; sum_other_d += row['기타공제']; sum_deduct_tot += emp_deduction_total; sum_net += net_pay
            sum_b_nat += row['국민연금(사업자)']; sum_b_hea += row['건강보험(사업자)']; sum_b_long += row['장기요양(사업자)']; sum_b_emp += row['고용보험(사업자)']; sum_b_ind += row['산재보험(사업자)']; sum_b_tot += biz_deduction_total; sum_retire += row['퇴직적립금']

            payroll_html_rows += f"""
            <tr>
                <td>{row['No']}</td><td>{row['이름']}</td><td>{row['생년월일']}</td><td>{row['호봉']}</td>
                <td style="text-align:right;">{row['기본급']:,}</td>
                <td style="text-align:right;">{row['초과수당(승인)']:,}</td>
                <td style="text-align:right;">{row['가족수당']:,}</td>
                <td style="text-align:right;">{row['비과세']:,}</td>
                <td style="text-align:right; font-weight:bold;">{total_gross:,}</td>
                <td style="text-align:right;">{row['국민연금(본인)']:,}</td>
                <td style="text-align:right;">{row['건강보험(본인)']:,}</td>
                <td style="text-align:right;">{row['장기요양(본인)']:,}</td>
                <td style="text-align:right;">{row['고용보험(본인)']:,}</td>
                <td style="text-align:right;">{row['소득세']:,}</td>
                <td style="text-align:right;">{row['지방소득세']:,}</td>
                <td style="text-align:right; font-weight:bold;">{emp_deduction_total:,}</td>
                <td style="text-align:right; font-weight:bold; background-color:#fffae6;">{net_pay:,}</td>
                <td style="text-align:right;">{row['국민연금(사업자)']:,}</td>
                <td style="text-align:right;">{row['건강보험(사업자)']:,}</td>
                <td style="text-align:right;">{row['장기요양(사업자)']:,}</td>
                <td style="text-align:right;">{row['고용보험(사업자)']:,}</td>
                <td style="text-align:right;">{row['산재보험(사업자)']:,}</td>
                <td style="text-align:right; font-weight:bold;">{biz_deduction_total:,}</td>
                <td style="text-align:right;">{row['퇴직적립금']:,}</td>
            </tr>
            """

        summary_html_row = f"""
        <tr style="background-color: #e6f2ff; font-weight: bold;">
            <td colspan="4">합 계</td>
            <td style="text-align:right;">{sum_base:,}</td>
            <td style="text-align:right;">{sum_ot:,}</td>
            <td style="text-align:right;">{sum_family:,}</td>
            <td style="text-align:right;">{sum_nontax:,}</td>
            <td style="text-align:right;">{sum_gross:,}</td>
            <td style="text-align:right;">{sum_nat:,}</td>
            <td style="text-align:right;">{sum_hea:,}</td>
            <td style="text-align:right;">{sum_long:,}</td>
            <td style="text-align:right;">{sum_emp:,}</td>
            <td style="text-align:right;">{sum_inc:,}</td>
            <td style="text-align:right;">{sum_loc:,}</td>
            <td style="text-align:right;">{sum_deduct_tot:,}</td>
            <td style="text-align:right; background-color:#ffe680;">{sum_net:,}</td>
            <td style="text-align:right;">{sum_b_nat:,}</td>
            <td style="text-align:right;">{sum_b_hea:,}</td>
            <td style="text-align:right;">{sum_b_long:,}</td>
            <td style="text-align:right;">{sum_b_emp:,}</td>
            <td style="text-align:right;">{sum_b_ind:,}</td>
            <td style="text-align:right;">{sum_b_tot:,}</td>
            <td style="text-align:right;">{sum_retire:,}</td>
        </tr>
        """

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            edited_payroll.to_excel(writer, index=False, sheet_name=f"{pay_month}_급여대장")
        excel_data = output.getvalue()

        st.download_button(
            label="📥 통합 급여대장 엑셀 다운로드 (.xlsx)",
            data=excel_data, file_name=f"통합급여대장_{pay_month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 28px; float: left;">' if st.session_state.logo_b64 else ''

        payroll_template = f"""
        <div style="font-family: 'Malgun Gothic', sans-serif; font-size: 11px; width: 100%; overflow-x: auto;">
            {logo_html}
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 5px; clear: both;">
                <div style="font-size: 14px; font-weight: bold;">지급일 &nbsp;&nbsp;&nbsp; {pay_date}</div>
                <div>(단위: 원 / 원단위 절사)</div>
            </div>
            
            <table border="1" style="width: 100%; border-collapse: collapse; text-align: center;" cellpadding="3">
                <thead>
                    <tr style="background-color: #ffffcc;">
                        <th rowspan="3" style="width: 25px;">No</th>
                        <th rowspan="3" style="width: 50px;">이름</th>
                        <th rowspan="3" style="width: 65px;">생년월일</th>
                        <th rowspan="3" style="width: 40px;">호봉</th>
                        <th colspan="4">지급 내역</th>
                        <th rowspan="3">급여총액</th>
                        <th colspan="7">근로자 본인 부담금</th>
                        <th rowspan="3" style="background-color: #fff2cc;">실지급액</th>
                        <th colspan="6">사업자 부담 사회보험금</th>
                        <th rowspan="3">사업주부담<br>퇴직적립금</th>
                    </tr>
                    <tr style="background-color: #ffffcc;">
                        <th rowspan="2">기본급</th>
                        <th rowspan="2">초과수당</th>
                        <th rowspan="2">가족수당</th>
                        <th rowspan="2">비과세</th>
                        <th>국민</th><th>건강</th><th>장기요양</th><th>고용</th><th>소득세</th><th>지방세</th>
                        <th rowspan="2">공제합계</th>
                        <th>국민</th><th>건강</th><th>장기요양</th><th>고용</th><th>산재</th>
                        <th rowspan="2">사업자합계</th>
                    </tr>
                    <tr style="background-color: #ffffcc;">
                        <th>4.75%</th><th>3.595%</th><th>12.95%</th><th>0.90%</th><th>간이세액</th><th>10%</th>
                        <th>4.75%</th><th>3.595%</th><th>12.95%</th><th>1.15%</th><th>7.26%</th>
                    </tr>
                </thead>
                <tbody>
                    {summary_html_row}
                    {payroll_html_rows}
                </tbody>
            </table>
        </div>
        """
        st.components.v1.html(payroll_template, height=520, scrolling=True)

# -------------------------------------------------------------------
# TAB 7: 개별 급여명세서 인쇄
# -------------------------------------------------------------------
with tab7:
    st.header("📄 개별 급여명세서 인쇄")
    
    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()

    ot_res = supabase.table("overtime_records").select("*").eq("status", "승인").execute()
    df_ot = pd.DataFrame(ot_res.data) if ot_res.data else pd.DataFrame()

    if df_emp.empty:
        st.warning("등록된 직원이 없다.")
    else:
        col1, col2 = st.columns(2)
        with col1:
            pay_month_slip = st.date_input("명세서 지급 월 선택", datetime.now(), key="slip_month").strftime("%Y-%m")
            emp_slip_list = df_emp['emp_name'] + " (" + df_emp['position'] + " / " + df_emp['emp_id'] + ")"
            selected_slip_str = st.selectbox("직원 선택", emp_slip_list, key="slip_emp")
            selected_slip_id = selected_slip_str.split("/")[-1].replace(")", "").strip()
            emp = df_emp[df_emp['emp_id'] == selected_slip_id].iloc[0]

        current_hourly_wage = int(emp['hourly_wage'])

        adj_single_res = supabase.table("monthly_payroll_adjust").select("*").eq("pay_month", pay_month_slip).eq("emp_id", emp['emp_id']).execute()
        df_adj_single = pd.DataFrame(adj_single_res.data) if adj_single_res.data else pd.DataFrame()

        emp_ot = df_ot[(df_ot['emp_id'] == emp['emp_id']) & (df_ot['work_date'].str.startswith(pay_month_slip))] if not df_ot.empty else pd.DataFrame()
        weekday_ot_hours = emp_ot[emp_ot['work_type'].str.contains("평일", na=False)]['actual_duration_hours'].sum() if not emp_ot.empty else 0.0
        holiday_ot_hours = emp_ot[emp_ot['work_type'].str.contains("휴일", na=False)]['actual_duration_hours'].sum() if not emp_ot.empty else 0.0
        total_ot_hours = weekday_ot_hours + holiday_ot_hours
        calculated_ot_pay = int(emp_ot['actual_pay'].sum()) if not emp_ot.empty else 0

        if not df_adj_single.empty:
            adj = df_adj_single.iloc[0]
            base = adj['base_salary']
            ot_pay = adj['ot_pay'] if adj['ot_pay'] > 0 else calculated_ot_pay
            family = adj['family_allowance']
            non_tax = adj['non_taxable']
            other_allow = adj['other_allowance']
            emp_national = adj['national_pension']
            emp_health = adj['health_insurance']
            emp_longterm = adj['longterm_care']
            emp_employment = adj['employment_insurance']
            emp_income_tax = adj['income_tax']
            emp_local_tax = adj['local_tax']
            other_deduct = adj['other_deduction']
        else:
            ot_pay = calculated_ot_pay
            base = emp['base_salary']
            family = emp['family_allowance']
            non_tax = emp['non_taxable']
            other_allow = emp['other_allowance']
            other_deduct = emp['other_deduction']

            total_gross_tmp = truncate_ten(base + ot_pay + family + non_tax + other_allow)
            taxable_gross_tmp = total_gross_tmp - non_tax

            emp_national = truncate_ten(taxable_gross_tmp * 0.0475) if emp['is_national'] == 1 else 0
            emp_health = truncate_ten(taxable_gross_tmp * 0.03595) if emp['is_health'] == 1 else 0
            emp_longterm = truncate_ten(emp_health * 0.1295) if emp['is_health'] == 1 else 0
            emp_employment = truncate_ten(taxable_gross_tmp * 0.0090) if emp['is_employment'] == 1 else 0
            emp_income_tax = truncate_ten(taxable_gross_tmp * 0.03)
            emp_local_tax = truncate_ten(emp_income_tax * 0.10)

        total_gross = base + ot_pay + family + non_tax + other_allow
        emp_deduction_total = emp_national + emp_health + emp_longterm + emp_employment + emp_income_tax + emp_local_tax + other_deduct
        net_pay = total_gross - emp_deduction_total

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 32px; float: left;">' if st.session_state.logo_b64 else ''

        payslip_template = f"""
        <div style="text-align: right; margin-bottom: 10px;">
            <button onclick="window.print()" style="padding: 8px 16px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px;">🖨️ 급여명세서 인쇄하기</button>
        </div>
        <div style="border: 2px solid #000; padding: 30px; font-family: 'Malgun Gothic', sans-serif; max-width: 680px; margin: auto; background: #fff;">
            {logo_html}
            <h2 style="text-align: center; margin-top: 0; margin-bottom: 25px; font-size: 24px; text-decoration: underline; clear: both;">
                {pay_month_slip}월 급 여 명 세 서
            </h2>

            <table style="width: 100%; border-collapse: collapse; margin-bottom: 15px; font-size: 13px;" border="1">
                <tr>
                    <th style="padding: 6px; background: #f2f2f2; width: 15%;">성 명</th>
                    <td style="padding: 6px; width: 35%;">{emp['emp_name']} ({emp['position']})</td>
                    <th style="padding: 6px; background: #f2f2f2; width: 15%;">소 속</th>
                    <td style="padding: 6px; width: 35%;">{emp['dept']}</td>
                </tr>
                <tr>
                    <th style="padding: 6px; background: #f2f2f2;">사 번</th>
                    <td style="padding: 6px;">{emp['emp_id']}</td>
                    <th style="padding: 6px; background: #f2f2f2;">생년월일</th>
                    <td style="padding: 6px;">{emp['birth_date']}</td>
                </tr>
            </table>

            <table style="width: 100%; border-collapse: collapse; margin-bottom: 15px; font-size: 13px;" border="1">
                <tr style="background: #e6f2ff;">
                    <th style="padding: 8px; width: 50%;" colspan="2">지급 항목</th>
                    <th style="padding: 8px; width: 50%;" colspan="2">공제 항목</th>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">기본급</td>
                    <td style="padding: 6px; text-align: right;">{base:,} 원</td>
                    <td style="padding: 6px; background: #f9f9f9;">국민연금</td>
                    <td style="padding: 6px; text-align: right;">{emp_national:,} 원</td>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">시간외수당 ({total_ot_hours:.1f}h)</td>
                    <td style="padding: 6px; text-align: right;">{ot_pay:,} 원</td>
                    <td style="padding: 6px; background: #f9f9f9;">건강보험</td>
                    <td style="padding: 6px; text-align: right;">{emp_health:,} 원</td>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">가족수당</td>
                    <td style="padding: 6px; text-align: right;">{family:,} 원</td>
                    <td style="padding: 6px; background: #f9f9f9;">장기요양보험</td>
                    <td style="padding: 6px; text-align: right;">{emp_longterm:,} 원</td>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">비과세 식대/보육</td>
                    <td style="padding: 6px; text-align: right;">{non_tax:,} 원</td>
                    <td style="padding: 6px; background: #f9f9f9;">고용보험</td>
                    <td style="padding: 6px; text-align: right;">{emp_employment:,} 원</td>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">기타수당</td>
                    <td style="padding: 6px; text-align: right;">{other_allow:,} 원</td>
                    <td style="padding: 6px; background: #f9f9f9;">소득세 / 지방소득세</td>
                    <td style="padding: 6px; text-align: right;">{(emp_income_tax + emp_local_tax):,} 원</td>
                </tr>
                <tr>
                    <td style="padding: 6px; background: #f9f9f9;">-</td>
                    <td style="padding: 6px; text-align: right;">-</td>
                    <td style="padding: 6px; background: #f9f9f9;">기타공제</td>
                    <td style="padding: 6px; text-align: right;">{other_deduct:,} 원</td>
                </tr>
                <tr style="font-weight: bold; background: #f2f2f2;">
                    <td style="padding: 8px;">지급액 계</td>
                    <td style="padding: 8px; text-align: right;">{total_gross:,} 원</td>
                    <td style="padding: 8px;">공제액 계</td>
                    <td style="padding: 8px; text-align: right;">{emp_deduction_total:,} 원</td>
                </tr>
            </table>

            <div style="border: 2px solid #333; padding: 10px; text-align: center; background: #fffde7; margin-bottom: 20px;">
                <span style="font-size: 16px; font-weight: bold;">실지급액 : {net_pay:,} 원</span>
            </div>

            <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px; text-align: center; font-size: 12px;" border="1">
                <tr style="background-color: #f2f2f2; height: 28px;">
                    <th style="width: 25%;">통상시급</th>
                    <th style="width: 25%;">시간외·연장근로시간</th>
                    <th style="width: 25%;">휴일근로시간</th>
                    <th style="width: 25%;">야간근로시간</th>
                </tr>
                <tr style="height: 32px;">
                    <td><b>{current_hourly_wage:,} 원</b></td>
                    <td>{weekday_ot_hours if weekday_ot_hours > 0 else '-'}</td>
                    <td>{holiday_ot_hours if holiday_ot_hours > 0 else '-'}</td>
                    <td>-</td>
                </tr>
            </table>

            <div style="font-size: 13px; font-weight: bold; text-align: center; margin-bottom: 8px; background-color: #f2f2f2; padding: 5px; border: 1px solid #000;">
                계 산 방 법
            </div>
            <table style="width: 100%; border-collapse: collapse; font-size: 11px;" border="1">
                <tr style="background-color: #f9f9f9; text-align: center; height: 25px;">
                    <th style="width: 35%;">구분</th>
                    <th style="width: 65%;">산출식 또는 산출방법</th>
                </tr>
                <tr style="height: 24px;">
                    <td style="text-align: center;">일할급여 계산기준<br>(중간 입사/퇴사 시)</td>
                    <td style="padding-left: 10px;">기본급 ÷ 해당 월일수 × 근무일수</td>
                </tr>
                <tr style="height: 24px;">
                    <td style="text-align: center;">연장근로수당</td>
                    <td style="padding-left: 10px;">연장근로시간수 × 통상시급({current_hourly_wage:,}원) × 1.5</td>
                </tr>
                <tr style="height: 24px;">
                    <td style="text-align: center;">휴일근로수당</td>
                    <td style="padding-left: 10px;">휴일근로시간수 × 통상시급({current_hourly_wage:,}원) × 1.5</td>
                </tr>
                <tr style="height: 24px;">
                    <td style="text-align: center;">야간근로수당</td>
                    <td style="padding-left: 10px;">야간근로시간수 × 통상시급({current_hourly_wage:,}원) × 1.5</td>
                </tr>
            </table>
            <p style="font-size: 10px; color: #333; margin-top: 5px; margin-bottom: 15px;">
                * 통상시급 : 정기적이고 일률적으로 지급하는 급여 ÷ 소정근로시간(209시간)
            </p>

            <p style="text-align: center; margin-top: 20px; font-size: 12px; color: #444;">귀하의 노고에 진심으로 감사드립니다.</p>
        </div>
        """
        st.components.v1.html(payslip_template, height=890, scrolling=True)

# -------------------------------------------------------------------
# TAB 8: 통합 급여대장 인쇄
# -------------------------------------------------------------------
with tab8:
    st.header("🖨️ 통합 급여대장 인쇄")
    
    pay_date_print = st.date_input("출력할 급여 지급일 선택", datetime.now(), key="payroll_print_date")
    pay_month_print = pay_date_print.strftime("%Y-%m")

    emp_res = supabase.table("employees").select("*").execute()
    df_emp = pd.DataFrame(emp_res.data) if emp_res.data else pd.DataFrame()

    ot_res = supabase.table("overtime_records").select("*").eq("status", "승인").execute()
    df_ot = pd.DataFrame(ot_res.data) if ot_res.data else pd.DataFrame()

    adj_res = supabase.table("monthly_payroll_adjust").select("*").eq("pay_month", pay_month_print).execute()
    df_adjust = pd.DataFrame(adj_res.data) if adj_res.data else pd.DataFrame()

    if df_emp.empty:
        st.warning("등록된 직원이 없다.")
    else:
        payroll_print_rows = ""
        no = 1

        sum_base = sum_ot = sum_family = sum_nontax = sum_gross = 0
        sum_nat = sum_hea = sum_long = sum_emp = sum_inc = sum_loc = sum_other_d = sum_deduct_tot = sum_net = 0
        sum_b_nat = sum_b_hea = sum_b_long = sum_b_emp = sum_b_ind = sum_b_tot = sum_retire = 0

        for idx, emp in df_emp.iterrows():
            adj_match = df_adjust[df_adjust['emp_id'] == emp['emp_id']] if not df_adjust.empty else pd.DataFrame()
            emp_ot = df_ot[(df_ot['emp_id'] == emp['emp_id']) & (df_ot['work_date'].str.startswith(pay_month_print))] if not df_ot.empty else pd.DataFrame()
            calculated_ot_pay = int(emp_ot['actual_pay'].sum()) if not emp_ot.empty else 0

            if not adj_match.empty:
                adj = adj_match.iloc[0]
                base = adj['base_salary']
                ot_pay = adj['ot_pay'] if adj['ot_pay'] > 0 else calculated_ot_pay
                family = adj['family_allowance']
                non_tax = adj['non_taxable']
                other_allow = adj['other_allowance']
                emp_national = adj['national_pension']
                emp_health = adj['health_insurance']
                emp_longterm = adj['longterm_care']
                emp_employment = adj['employment_insurance']
                emp_income_tax = adj['income_tax']
                emp_local_tax = adj['local_tax']
                other_deduct = adj['other_deduction']
            else:
                ot_pay = calculated_ot_pay
                base = emp['base_salary']
                family = emp['family_allowance']
                non_tax = emp['non_taxable']
                other_allow = emp['other_allowance']
                other_deduct = emp['other_deduction']

                total_gross_calc = truncate_ten(base + ot_pay + family + non_tax + other_allow)
                taxable_gross_calc = total_gross_calc - non_tax

                emp_national = truncate_ten(taxable_gross_calc * 0.0475) if emp.get('is_national', 1) == 1 else 0
                emp_health = truncate_ten(taxable_gross_calc * 0.03595) if emp.get('is_health', 1) == 1 else 0
                emp_longterm = truncate_ten(emp_health * 0.1295) if emp.get('is_health', 1) == 1 else 0
                emp_employment = truncate_ten(taxable_gross_calc * 0.0090) if emp.get('is_employment', 1) == 1 else 0
                emp_income_tax = truncate_ten(taxable_gross_calc * 0.03)
                emp_local_tax = truncate_ten(emp_income_tax * 0.10)

            tot_g = base + ot_pay + family + non_tax + other_allow
            taxable_gross = tot_g - non_tax
            emp_deduction_total = emp_national + emp_health + emp_longterm + emp_employment + emp_income_tax + emp_local_tax + other_deduct
            net_pay = tot_g - emp_deduction_total

            biz_national = truncate_ten(taxable_gross * 0.0475) if emp.get('is_national', 1) == 1 else 0
            biz_health = truncate_ten(taxable_gross * 0.03595) if emp.get('is_health', 1) == 1 else 0
            biz_longterm = truncate_ten(biz_health * 0.1295) if emp.get('is_health', 1) == 1 else 0
            biz_employment = truncate_ten(taxable_gross * 0.0115) if emp.get('is_employment', 1) == 1 else 0
            biz_industrial = truncate_ten(taxable_gross * 0.0726) if emp.get('is_industrial', 1) == 1 else 0
            biz_deduction_total = biz_national + biz_health + biz_longterm + biz_employment + biz_industrial
            retirement_accrual = truncate_ten(tot_g / 12)

            sum_base += base; sum_ot += ot_pay; sum_family += family; sum_nontax += non_tax; sum_gross += tot_g
            sum_nat += emp_national; sum_hea += emp_health; sum_long += emp_longterm; sum_emp += emp_employment
            sum_inc += emp_income_tax; sum_loc += emp_local_tax; sum_other_d += other_deduct; sum_deduct_tot += emp_deduction_total; sum_net += net_pay
            sum_b_nat += biz_national; sum_b_hea += biz_health; sum_b_long += biz_longterm; sum_b_emp += biz_employment; sum_b_ind += biz_industrial; sum_b_tot += biz_deduction_total; sum_retire += retirement_accrual

            payroll_print_rows += f"""
            <tr style="height: 26px;">
                <td>{no}</td><td>{emp['emp_name']}</td><td>{emp['birth_date']}</td><td>{emp['hobong']}</td>
                <td style="text-align:right;">{base:,}</td>
                <td style="text-align:right;">{ot_pay:,}</td>
                <td style="text-align:right;">{family:,}</td>
                <td style="text-align:right;">{non_tax:,}</td>
                <td style="text-align:right; font-weight:bold;">{tot_g:,}</td>
                <td style="text-align:right;">{emp_national:,}</td>
                <td style="text-align:right;">{emp_health:,}</td>
                <td style="text-align:right;">{emp_longterm:,}</td>
                <td style="text-align:right;">{emp_employment:,}</td>
                <td style="text-align:right;">{emp_income_tax:,}</td>
                <td style="text-align:right;">{emp_local_tax:,}</td>
                <td style="text-align:right; font-weight:bold;">{emp_deduction_total:,}</td>
                <td style="text-align:right; font-weight:bold; background-color:#fffae6;">{net_pay:,}</td>
                <td style="text-align:right;">{biz_national:,}</td>
                <td style="text-align:right;">{biz_health:,}</td>
                <td style="text-align:right;">{biz_longterm:,}</td>
                <td style="text-align:right;">{biz_employment:,}</td>
                <td style="text-align:right;">{biz_industrial:,}</td>
                <td style="text-align:right; font-weight:bold;">{biz_deduction_total:,}</td>
                <td style="text-align:right;">{retirement_accrual:,}</td>
            </tr>
            """
            no += 1

        summary_print_row = f"""
        <tr style="background-color: #e6f2ff; font-weight: bold; height: 28px;">
            <td colspan="4">합 계</td>
            <td style="text-align:right;">{sum_base:,}</td>
            <td style="text-align:right;">{sum_ot:,}</td>
            <td style="text-align:right;">{sum_family:,}</td>
            <td style="text-align:right;">{sum_nontax:,}</td>
            <td style="text-align:right;">{sum_gross:,}</td>
            <td style="text-align:right;">{sum_nat:,}</td>
            <td style="text-align:right;">{sum_hea:,}</td>
            <td style="text-align:right;">{sum_long:,}</td>
            <td style="text-align:right;">{sum_emp:,}</td>
            <td style="text-align:right;">{sum_inc:,}</td>
            <td style="text-align:right;">{sum_loc:,}</td>
            <td style="text-align:right;">{sum_deduct_tot:,}</td>
            <td style="text-align:right; background-color:#ffe680;">{sum_net:,}</td>
            <td style="text-align:right;">{sum_b_nat:,}</td>
            <td style="text-align:right;">{sum_b_hea:,}</td>
            <td style="text-align:right;">{sum_b_long:,}</td>
            <td style="text-align:right;">{sum_b_emp:,}</td>
            <td style="text-align:right;">{sum_b_ind:,}</td>
            <td style="text-align:right;">{sum_b_tot:,}</td>
            <td style="text-align:right;">{sum_retire:,}</td>
        </tr>
        """

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 28px; float: left;">' if st.session_state.logo_b64 else ''

        payroll_print_template = f"""
        <div style="text-align: right; margin-bottom: 10px;">
            <button onclick="window.print()" style="padding: 8px 16px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px;">🖨️ 급여대장 인쇄하기</button>
        </div>
        <div style="border: 2px solid #000; padding: 20px; font-family: 'Malgun Gothic', sans-serif; background: #fff; width: 100%; box-sizing: border-box;">
            {logo_html}
            <div style="display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 15px; clear: both;">
                <h2 style="margin: 0; padding-top: 5px; font-size: 22px; text-decoration: underline;">{pay_month_print}월 통합 급여대장</h2>
                <table style="border-collapse: collapse; text-align: center; font-size: 11px; width: 210px;" border="1">
                    <tr style="height: 18px; background-color: #f2f2f2;">
                        <th rowspan="2" style="width: 25px; background-color: #e6e6e6;">결<br>재</th>
                        <th style="width: 60px;">담 당</th>
                        <th style="width: 60px;">대 리</th>
                        <th style="width: 65px;">센터장</th>
                    </tr>
                    <tr style="height: 40px;">
                        <td></td><td></td><td></td>
                    </tr>
                </table>
            </div>

            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; font-size: 11px;">
                <div style="font-size: 12px; font-weight: bold;">지급일자: {pay_date_print}</div>
                <div>(단위: 원 / 원단위 절사)</div>
            </div>

            <table border="1" style="width: 100%; border-collapse: collapse; text-align: center; font-size: 10px;" cellpadding="2">
                <thead>
                    <tr style="background-color: #ffffcc;">
                        <th rowspan="3" style="width: 25px;">No</th>
                        <th rowspan="3" style="width: 50px;">이름</th>
                        <th rowspan="3" style="width: 65px;">생년월일</th>
                        <th rowspan="3" style="width: 40px;">호봉</th>
                        <th colspan="4">지급 내역</th>
                        <th rowspan="3">급여총액</th>
                        <th colspan="7">근로자 본인 부담금</th>
                        <th rowspan="3" style="background-color: #fff2cc;">실지급액</th>
                        <th colspan="6">사업자 부담 사회보험금</th>
                        <th rowspan="3">사업주부담<br>퇴직적립금</th>
                    </tr>
                    <tr style="background-color: #ffffcc;">
                        <th rowspan="2">기본급</th>
                        <th rowspan="2">초과수당</th>
                        <th rowspan="2">가족수당</th>
                        <th rowspan="2">비과세</th>
                        <th>국민</th><th>건강</th><th>장기요양</th><th>고용</th><th>소득세</th><th>지방세</th>
                        <th rowspan="2">공제합계</th>
                        <th>국민</th><th>건강</th><th>장기요양</th><th>고용</th><th>산재</th>
                        <th rowspan="2">사업자합계</th>
                    </tr>
                    <tr style="background-color: #ffffcc;">
                        <th>4.75%</th><th>3.595%</th><th>12.95%</th><th>0.90%</th><th>간이세액</th><th>10%</th>
                        <th>4.75%</th><th>3.595%</th><th>12.95%</th><th>1.15%</th><th>7.26%</th>
                    </tr>
                </thead>
                <tbody>
                    {summary_print_row}
                    {payroll_print_rows}
                </tbody>
            </table>
        </div>
        """
        st.components.v1.html(payroll_print_template, height=600, scrolling=True)

# -------------------------------------------------------------------
# TAB 9: 월별 급여대장 총괄표
# -------------------------------------------------------------------
with tab9:
    st.header("📑 월별 급여대장 총괄표 (12개월 누적 요약)")
    
    c_y9 = st.selectbox("조회 연도 선택", range(datetime.now().year - 2, datetime.now().year + 3), index=2, key="annual_summary_year")
    
    emp_all_res = supabase.table("employees").select("*").execute()
    df_emp_all = pd.DataFrame(emp_all_res.data) if emp_all_res.data else pd.DataFrame()

    ot_all_res = supabase.table("overtime_records").select("*").eq("status", "승인").execute()
    df_ot_all = pd.DataFrame(ot_all_res.data) if ot_all_res.data else pd.DataFrame()

    adj_all_res = supabase.table("monthly_payroll_adjust").select("*").execute()
    df_adj_all = pd.DataFrame(adj_all_res.data) if adj_all_res.data else pd.DataFrame()

    if df_emp_all.empty:
        st.warning("등록된 직원 정보가 없다.")
    else:
        monthly_summary_rows = []
        tot_ann_count = tot_ann_base = tot_ann_ot_hours = tot_ann_ot = tot_ann_fam = tot_ann_nontax = tot_ann_other_a = 0
        tot_ann_gross = tot_ann_nat = tot_ann_hea = tot_ann_long = tot_ann_emp = tot_ann_inc = tot_ann_loc = 0
        tot_ann_other_d = tot_ann_deduct = tot_ann_net = tot_ann_retire = 0

        for m in range(1, 13):
            m_str = f"{c_y9}-{m:02d}"
            m_emp_count = len(df_emp_all)
            m_base = m_ot = m_fam = m_nontax = m_other_a = 0
            m_nat = m_hea = m_long = m_emp = m_inc = m_loc = m_other_d = 0
            m_ot_hours = 0.0

            m_ot_records = df_ot_all[df_ot_all['work_date'].str.startswith(m_str)] if not df_ot_all.empty else pd.DataFrame()
            if not m_ot_records.empty:
                m_ot_hours = m_ot_records['actual_duration_hours'].sum()

            for _, emp in df_emp_all.iterrows():
                adj_m = df_adj_all[(df_adj_all['pay_month'] == m_str) & (df_adj_all['emp_id'] == emp['emp_id'])] if not df_adj_all.empty else pd.DataFrame()
                
                emp_ot = df_ot_all[(df_ot_all['emp_id'] == emp['emp_id']) & (df_ot_all['work_date'].str.startswith(m_str))] if not df_ot_all.empty else pd.DataFrame()
                calc_ot = int(emp_ot['actual_pay'].sum()) if not emp_ot.empty else 0

                if not adj_m.empty:
                    adj = adj_m.iloc[0]
                    base = adj['base_salary']
                    ot = adj['ot_pay'] if adj['ot_pay'] > 0 else calc_ot
                    fam = adj['family_allowance']
                    nontax = adj['non_taxable']
                    other_a = adj['other_allowance']
                    nat = adj['national_pension']
                    hea = adj['health_insurance']
                    lng = adj['longterm_care']
                    e_emp = adj['employment_insurance']
                    inc = adj['income_tax']
                    loc = adj['local_tax']
                    other_d = adj['other_deduction']
                else:
                    ot = calc_ot
                    base = emp['base_salary']
                    fam = emp['family_allowance']
                    nontax = emp['non_taxable']
                    other_a = emp['other_allowance']
                    other_d = emp['other_deduction']

                    tot_g_tmp = truncate_ten(base + ot + fam + nontax + other_a)
                    taxable_tmp = tot_g_tmp - nontax

                    nat = truncate_ten(taxable_tmp * 0.0475) if emp.get('is_national', 1) == 1 else 0
                    hea = truncate_ten(taxable_tmp * 0.03595) if emp.get('is_health', 1) == 1 else 0
                    lng = truncate_ten(hea * 0.1295) if emp.get('is_health', 1) == 1 else 0
                    e_emp = truncate_ten(taxable_tmp * 0.0090) if emp.get('is_employment', 1) == 1 else 0
                    inc = truncate_ten(taxable_tmp * 0.03)
                    loc = truncate_ten(inc * 0.10)

                m_base += base; m_ot += ot; m_fam += fam; m_nontax += nontax; m_other_a += other_a
                m_nat += nat; m_hea += hea; m_long += lng; m_emp += e_emp
                m_inc += inc; m_loc += loc; m_other_d += other_d

            m_gross = m_base + m_ot + m_fam + m_nontax + m_other_a
            m_deduct = m_nat + m_hea + m_long + m_emp + m_inc + m_loc + m_other_d
            m_net = m_gross - m_deduct
            m_retire = truncate_ten(m_gross / 12)

            tot_ann_base += m_base; tot_ann_ot_hours += m_ot_hours; tot_ann_ot += m_ot; tot_ann_fam += m_fam; tot_ann_nontax += m_nontax; tot_ann_other_a += m_other_a
            tot_ann_gross += m_gross; tot_ann_nat += m_nat; tot_ann_hea += m_hea; tot_ann_long += m_long; tot_ann_emp += m_emp
            tot_ann_inc += m_inc; tot_ann_loc += m_loc; tot_ann_other_d += m_other_d; tot_ann_deduct += m_deduct; tot_ann_net += m_net
            tot_ann_retire += m_retire

            monthly_summary_rows.append(f"""
            <tr style="height: 26px;">
                <td><b>{m}월</b> ({m_str})</td>
                <td>{m_emp_count} 명</td>
                <td style="text-align:right;">{m_base:,}</td>
                <td style="text-align:right; background-color:#f0f8ff;"><b>{m_ot_hours:.1f} 시간</b></td>
                <td style="text-align:right;">{m_ot:,}</td>
                <td style="text-align:right;">{m_fam:,}</td>
                <td style="text-align:right;">{m_nontax:,}</td>
                <td style="text-align:right; font-weight:bold; background-color:#f9f9f9;">{m_gross:,}</td>
                <td style="text-align:right;">{m_nat:,}</td>
                <td style="text-align:right;">{m_hea:,}</td>
                <td style="text-align:right;">{m_long:,}</td>
                <td style="text-align:right;">{m_emp:,}</td>
                <td style="text-align:right;">{m_inc:,}</td>
                <td style="text-align:right;">{m_loc:,}</td>
                <td style="text-align:right; font-weight:bold; background-color:#f9f9f9;">{m_deduct:,}</td>
                <td style="text-align:right; font-weight:bold; background-color:#fffae6;">{m_net:,}</td>
                <td style="text-align:right;">{m_retire:,}</td>
            </tr>
            """)

        ann_sum_html_row = f"""
        <tr style="background-color: #e6f2ff; font-weight: bold; height: 30px;">
            <td colspan="2">연간 누적 합계</td>
            <td style="text-align:right;">{tot_ann_base:,}</td>
            <td style="text-align:right; background-color:#d0e8ff;">{tot_ann_ot_hours:.1f} 시간</td>
            <td style="text-align:right;">{tot_ann_ot:,}</td>
            <td style="text-align:right;">{tot_ann_fam:,}</td>
            <td style="text-align:right;">{tot_ann_nontax:,}</td>
            <td style="text-align:right;">{tot_ann_gross:,}</td>
            <td style="text-align:right;">{tot_ann_nat:,}</td>
            <td style="text-align:right;">{tot_ann_hea:,}</td>
            <td style="text-align:right;">{tot_ann_long:,}</td>
            <td style="text-align:right;">{tot_ann_emp:,}</td>
            <td style="text-align:right;">{tot_ann_inc:,}</td>
            <td style="text-align:right;">{tot_ann_loc:,}</td>
            <td style="text-align:right;">{tot_ann_deduct:,}</td>
            <td style="text-align:right; background-color:#ffe680;">{tot_ann_net:,}</td>
            <td style="text-align:right;">{tot_ann_retire:,}</td>
        </tr>
        """

        logo_html = f'<img src="data:image/png;base64,{st.session_state.logo_b64}" style="max-height: 28px; float: left;">' if st.session_state.logo_b64 else ''

        annual_summary_template = f"""
        <div style="text-align: right; margin-bottom: 10px;">
            <button onclick="window.print()" style="padding: 8px 16px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px;">🖨️ 총괄표 인쇄하기</button>
        </div>
        <div style="border: 2px solid #000; padding: 25px; font-family: 'Malgun Gothic', sans-serif; background: #fff; width: 100%; box-sizing: border-box;">
            {logo_html}
            <div style="display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px; clear: both;">
                <h2 style="margin: 0; padding-top: 5px; font-size: 22px; text-decoration: underline;">{c_y9}년도 월별 급여대장 총괄표</h2>
                <table style="border-collapse: collapse; text-align: center; font-size: 11px; width: 210px;" border="1">
                    <tr style="height: 18px; background-color: #f2f2f2;">
                        <th rowspan="2" style="width: 25px; background-color: #e6e6e6;">결<br>재</th>
                        <th style="width: 60px;">담 당</th>
                        <th style="width: 60px;">대 리</th>
                        <th style="width: 65px;">센터장</th>
                    </tr>
                    <tr style="height: 40px;">
                        <td></td><td></td><td></td>
                    </tr>
                </table>
            </div>

            <div style="text-align: right; margin-bottom: 8px; font-size: 11px;">(단위: 원 / 원단위 절사)</div>

            <table border="1" style="width: 100%; border-collapse: collapse; text-align: center; font-size: 11px;" cellpadding="3">
                <thead>
                    <tr style="background-color: #ffffcc; height: 32px;">
                        <th style="width: 80px;">지급월</th>
                        <th style="width: 45px;">인원</th>
                        <th>기본급</th>
                        <th style="background-color: #e6f2ff;">승인 초과시간</th>
                        <th>초과수당</th>
                        <th>가족수당</th>
                        <th>비과세</th>
                        <th style="background-color: #fff2cc;">급여총액</th>
                        <th>국민연금</th>
                        <th>건강보험</th>
                        <th>장기요양</th>
                        <th>고용보험</th>
                        <th>소득세</th>
                        <th>지방세</th>
                        <th style="background-color: #fff2cc;">공제합계</th>
                        <th style="background-color: #ffe680;">실지급액</th>
                        <th>퇴직적립금</th>
                    </tr>
                </thead>
                <tbody>
                    {ann_sum_html_row}
                    {''.join(monthly_summary_rows)}
                    {ann_sum_html_row}
                </tbody>
            </table>
        </div>
        """
        st.components.v1.html(annual_summary_template, height=650, scrolling=True)
